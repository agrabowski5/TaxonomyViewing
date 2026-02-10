"""
Generate tree and lookup JSON files for all 6 taxonomies:
  HS (international), CPC, CN (EU), HTS (US), Canadian Customs Tariff, UNSPSC

Reads raw data from ../raw-data/ and writes JSON to ../app/public/data/
"""

import csv
import json
import os
import re
import sys

RAW_DIR = os.path.join(os.path.dirname(__file__), '..', 'raw-data')
OUT_DIR = os.path.join(os.path.dirname(__file__), '..', 'app', 'public', 'data')

os.makedirs(OUT_DIR, exist_ok=True)


def write_json(filename, data):
    path = os.path.join(OUT_DIR, filename)
    with open(path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False)
    size = os.path.getsize(path)
    print(f"  Wrote {filename}: {size:,} bytes")


# ─────────────────────────────────────────────
# 1. HS (Harmonized System)
# ─────────────────────────────────────────────
def generate_hs():
    print("\n=== Generating HS data ===")

    # Load sections
    sections = {}
    with open(os.path.join(RAW_DIR, 'hs-sections.csv'), 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            sections[row['section'].strip()] = row['name'].strip()

    # Load codes
    codes = []
    with open(os.path.join(RAW_DIR, 'hs-codes.csv'), 'r', encoding='utf-8') as f:
        reader = csv.DictReader(f)
        for row in reader:
            codes.append(row)

    # Build chapter -> section mapping
    chapter_to_section = {}
    for code in codes:
        if code['level'] == '2':  # chapter
            chapter_to_section[code['hscode']] = code['section']

    # Build tree
    tree = []
    section_nodes = {}  # section roman -> node
    chapter_nodes = {}  # chapter code -> node
    heading_nodes = {}  # heading code -> node

    # Create section nodes
    for section_roman, section_name in sections.items():
        node = {
            'id': f'hs-section-{section_roman}',
            'code': section_roman,
            'name': section_name.title() if section_name[0].islower() else section_name,
            'type': 'section',
            'children': []
        }
        section_nodes[section_roman] = node
        tree.append(node)

    # Process codes
    for code in codes:
        level = code['level']
        hscode = code['hscode']
        desc = code['description']
        section = code['section']
        parent = code['parent']

        if level == '2':  # chapter
            node = {
                'id': f'hs-{hscode}',
                'code': hscode,
                'name': desc,
                'type': 'chapter',
                'children': []
            }
            chapter_nodes[hscode] = node
            if section in section_nodes:
                section_nodes[section]['children'].append(node)
        elif level == '4':  # heading
            node = {
                'id': f'hs-{hscode}',
                'code': hscode,
                'name': desc,
                'type': 'heading',
                'children': []
            }
            heading_nodes[hscode] = node
            if parent in chapter_nodes:
                chapter_nodes[parent]['children'].append(node)
        elif level == '6' or level == '5':  # subheading
            node = {
                'id': f'hs-{hscode}',
                'code': hscode,
                'name': desc,
                'type': 'subheading'
            }
            if parent in heading_nodes:
                heading_nodes[parent]['children'].append(node)

    # Remove empty children arrays from leaf nodes
    def clean_tree(nodes):
        for node in nodes:
            if 'children' in node:
                if len(node['children']) == 0:
                    del node['children']
                else:
                    clean_tree(node['children'])

    clean_tree(tree)

    # Build lookup
    lookup = {}
    for code in codes:
        hscode = code['hscode']
        section = code['section']
        level_map = {'2': 'chapter', '4': 'heading', '5': 'subheading', '6': 'subheading'}
        lookup[hscode] = {
            'code': hscode,
            'description': code['description'],
            'section': section,
            'sectionName': sections.get(section, ''),
            'level': int(code['level']),
            'type': level_map.get(code['level'], 'unknown')
        }

    write_json('hs-tree.json', tree)
    write_json('hs-lookup.json', lookup)
    print(f"  HS: {len(tree)} sections, {len(codes)} total codes")


# ─────────────────────────────────────────────
# 2. CPC (Central Product Classification)
# ─────────────────────────────────────────────
def generate_cpc():
    print("\n=== Generating CPC data ===")

    codes = []
    with open(os.path.join(RAW_DIR, 'cpc-structure.txt'), 'r', encoding='latin-1') as f:
        reader = csv.reader(f)
        next(reader)  # skip header
        for row in reader:
            if len(row) >= 2:
                codes.append({'code': row[0].strip(), 'title': row[1].strip()})

    level_names = {1: 'section', 2: 'division', 3: 'group', 4: 'class', 5: 'subclass'}

    # Build tree
    tree = []
    nodes_by_code = {}

    for entry in codes:
        code = entry['code']
        title = entry['title']
        level = len(code)
        type_name = level_names.get(level, 'item')

        node = {
            'id': f'cpc-{code}',
            'code': code,
            'name': title,
            'type': type_name,
        }

        if level < 5:
            node['children'] = []

        nodes_by_code[code] = node

        if level == 1:
            tree.append(node)
        else:
            parent_code = code[:-1]
            if parent_code in nodes_by_code:
                parent = nodes_by_code[parent_code]
                if 'children' not in parent:
                    parent['children'] = []
                parent['children'].append(node)

    # Clean empty children
    def clean_tree(nodes):
        for node in nodes:
            if 'children' in node:
                if len(node['children']) == 0:
                    del node['children']
                else:
                    clean_tree(node['children'])

    clean_tree(tree)

    # Build lookup
    lookup = {}
    for entry in codes:
        code = entry['code']
        level = len(code)
        # Find section
        section_code = code[0] if code else ''
        section_name = ''
        if section_code in nodes_by_code:
            section_name = nodes_by_code[section_code]['name']

        lookup[code] = {
            'code': code,
            'description': entry['title'],
            'section': section_code,
            'sectionName': section_name,
            'level': level,
            'type': level_names.get(level, 'item')
        }

    write_json('cpc-tree.json', tree)
    write_json('cpc-lookup.json', lookup)
    print(f"  CPC: {len(tree)} sections, {len(codes)} total codes")


# ─────────────────────────────────────────────
# 3. CN (EU Combined Nomenclature)
# ─────────────────────────────────────────────
def generate_cn():
    print("\n=== Generating CN data ===")
    try:
        import openpyxl
    except ImportError:
        print("  ERROR: openpyxl not installed. Run: pip install openpyxl")
        return

    wb = openpyxl.load_workbook(os.path.join(RAW_DIR, 'cn-2025.xlsx'))
    ws = wb[wb.sheetnames[0]]

    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):  # skip header
        cnkey, cn, dashes, dm, su = row[0], row[1], row[2], row[3], row[4] if len(row) > 4 else None
        if dm:
            rows.append({
                'cnkey': str(cnkey) if cnkey else '',
                'cn': str(cn).strip() if cn else '',
                'dashes': str(dashes).strip() if dashes else '',
                'description': str(dm).strip(),
                'su': str(su).strip() if su else ''
            })

    # Parse CN codes and build tree
    tree = []
    section_nodes = {}
    chapter_nodes = {}
    heading_nodes = {}
    subheading_nodes = {}
    cn8_nodes = {}

    # HS section mapping (chapter -> section)
    hs_sections = {}
    with open(os.path.join(RAW_DIR, 'hs-sections.csv'), 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            hs_sections[row['section'].strip()] = row['name'].strip()

    current_section = None

    for row in rows:
        cn_code = row['cn']
        desc = row['description']
        dashes = row['dashes']

        # Skip the overall title row
        if cn_code == '' and dashes == '' and 'COMBINED NOMENCLATURE' in desc:
            continue

        # Section (Roman numeral)
        if cn_code and re.match(r'^[IVX]+$', cn_code):
            current_section = cn_code
            section_name = desc
            # Clean up section name (remove "SECTION I - " prefix)
            section_name = re.sub(r'^SECTION\s+[IVX]+\s*[-–]\s*', '', section_name)
            node = {
                'id': f'cn-section-{cn_code}',
                'code': cn_code,
                'name': section_name,
                'type': 'section',
                'children': []
            }
            section_nodes[cn_code] = node
            tree.append(node)
            continue

        # Chapter (2-digit)
        if cn_code and re.match(r'^\d{2}$', cn_code):
            chapter_name = desc
            # Clean up "CHAPTER X - " prefix
            chapter_name = re.sub(r'^CHAPTER\s+\d+\s*[-–]\s*', '', chapter_name)
            node = {
                'id': f'cn-{cn_code}',
                'code': cn_code,
                'name': chapter_name,
                'type': 'chapter',
                'children': []
            }
            chapter_nodes[cn_code] = node
            if current_section and current_section in section_nodes:
                section_nodes[current_section]['children'].append(node)
            continue

        # Heading (4-digit)
        if cn_code and re.match(r'^\d{4}$', cn_code):
            node = {
                'id': f'cn-{cn_code}',
                'code': cn_code,
                'name': desc,
                'type': 'heading',
                'children': []
            }
            heading_nodes[cn_code] = node
            chapter = cn_code[:2]
            if chapter in chapter_nodes:
                chapter_nodes[chapter]['children'].append(node)
            continue

        # 6-digit subheading (format: "0101 21" or "0101 29")
        clean_cn = cn_code.replace(' ', '')
        if cn_code and re.match(r'^\d{4}\s+\d{2}$', cn_code):
            node = {
                'id': f'cn-{clean_cn}',
                'code': clean_cn,
                'name': desc,
                'type': 'subheading',
                'children': []
            }
            subheading_nodes[clean_cn] = node
            heading = clean_cn[:4]
            if heading in heading_nodes:
                heading_nodes[heading]['children'].append(node)
            continue

        # 8-digit CN code (format: "0101 21 00")
        if cn_code and re.match(r'^\d{4}\s+\d{2}\s+\d{2}$', cn_code):
            node = {
                'id': f'cn-{clean_cn}',
                'code': clean_cn,
                'name': desc,
                'type': 'cn8'
            }
            cn8_nodes[clean_cn] = node
            subheading = clean_cn[:6]
            if subheading in subheading_nodes:
                subheading_nodes[subheading]['children'].append(node)
            elif clean_cn[:4] in heading_nodes:
                heading_nodes[clean_cn[:4]]['children'].append(node)
            continue

    # Clean empty children
    def clean_tree(nodes):
        for node in nodes:
            if 'children' in node:
                if len(node['children']) == 0:
                    del node['children']
                else:
                    clean_tree(node['children'])

    clean_tree(tree)

    # Build lookup
    lookup = {}

    def add_to_lookup(nodes, section='', section_name=''):
        for node in nodes:
            if node['type'] == 'section':
                section = node['code']
                section_name = node['name']
            code = node['code']
            level_map = {'section': 1, 'chapter': 2, 'heading': 4, 'subheading': 6, 'cn8': 8}
            lookup[code] = {
                'code': code,
                'description': node['name'],
                'section': section,
                'sectionName': section_name,
                'level': level_map.get(node['type'], 0),
                'type': node['type']
            }
            if 'children' in node:
                add_to_lookup(node['children'], section, section_name)

    add_to_lookup(tree)

    write_json('cn-tree.json', tree)
    write_json('cn-lookup.json', lookup)
    count = sum(1 for _ in lookup)
    print(f"  CN: {len(tree)} sections, {count} total codes")


# ─────────────────────────────────────────────
# 4. HTS (US Harmonized Tariff Schedule)
# ─────────────────────────────────────────────
def generate_hts():
    print("\n=== Generating HTS data ===")

    with open(os.path.join(RAW_DIR, 'hts-2026.json'), 'r', encoding='utf-8') as f:
        raw = json.load(f)

    # Load HS sections for section-level grouping
    hs_sections = {}
    with open(os.path.join(RAW_DIR, 'hs-sections.csv'), 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            hs_sections[row['section'].strip()] = row['name'].strip()

    # HS chapter to section mapping
    chapter_to_section = {
        '01': 'I', '02': 'I', '03': 'I', '04': 'I', '05': 'I',
        '06': 'II', '07': 'II', '08': 'II', '09': 'II', '10': 'II', '11': 'II', '12': 'II', '13': 'II', '14': 'II',
        '15': 'III',
        '16': 'IV', '17': 'IV', '18': 'IV', '19': 'IV', '20': 'IV', '21': 'IV', '22': 'IV', '23': 'IV', '24': 'IV',
        '25': 'V', '26': 'V', '27': 'V',
        '28': 'VI', '29': 'VI', '30': 'VI', '31': 'VI', '32': 'VI', '33': 'VI', '34': 'VI', '35': 'VI', '36': 'VI', '37': 'VI', '38': 'VI',
        '39': 'VII', '40': 'VII',
        '41': 'VIII', '42': 'VIII', '43': 'VIII',
        '44': 'IX', '45': 'IX', '46': 'IX',
        '47': 'X', '48': 'X', '49': 'X',
        '50': 'XI', '51': 'XI', '52': 'XI', '53': 'XI', '54': 'XI', '55': 'XI', '56': 'XI', '57': 'XI', '58': 'XI', '59': 'XI', '60': 'XI', '61': 'XI', '62': 'XI', '63': 'XI',
        '64': 'XII', '65': 'XII', '66': 'XII', '67': 'XII',
        '68': 'XIII', '69': 'XIII', '70': 'XIII',
        '71': 'XIV',
        '72': 'XV', '73': 'XV', '74': 'XV', '75': 'XV', '76': 'XV', '78': 'XV', '79': 'XV', '80': 'XV', '81': 'XV', '82': 'XV', '83': 'XV',
        '84': 'XVI', '85': 'XVI',
        '86': 'XVII', '87': 'XVII', '88': 'XVII', '89': 'XVII',
        '90': 'XVIII', '91': 'XVIII', '92': 'XVIII',
        '93': 'XIX',
        '94': 'XX', '95': 'XX', '96': 'XX',
        '97': 'XXI',
        '98': 'XXII', '99': 'XXII',
    }

    # Build tree using indent-based hierarchy
    tree = []
    section_nodes = {}
    # Stack to track parent at each indent level
    stack = []  # list of (indent_level, node)

    # First pass: create section nodes
    for section_roman, section_name in hs_sections.items():
        node = {
            'id': f'hts-section-{section_roman}',
            'code': section_roman,
            'name': section_name.title() if section_name[0].islower() else section_name,
            'type': 'section',
            'children': []
        }
        section_nodes[section_roman] = node
        tree.append(node)

    # Add US-specific sections for chapters 98-99
    if 'XXII' not in section_nodes:
        node = {
            'id': 'hts-section-XXII',
            'code': 'XXII',
            'name': 'Special Classification Provisions',
            'type': 'section',
            'children': []
        }
        section_nodes['XXII'] = node
        tree.append(node)

    # Second pass: process HTS entries
    current_chapter = None
    chapter_nodes_map = {}

    for item in raw:
        htsno = item.get('htsno', '').strip()
        indent = int(item.get('indent', 0))
        desc = item.get('description', '').strip()
        superior = item.get('superior')

        if not desc:
            continue

        # Clean description - remove trailing colons
        clean_desc = desc.rstrip(':').strip()

        # Determine code and type
        if htsno:
            clean_code = htsno.replace('.', '')
            code_len = len(clean_code)
        else:
            clean_code = ''
            code_len = 0

        # Determine node type
        if code_len == 4:
            node_type = 'heading'
            current_chapter = clean_code[:2]
        elif code_len == 6:
            node_type = 'subheading'
        elif code_len == 8:
            node_type = 'tariff_8'
        elif code_len == 10:
            node_type = 'tariff_10'
        else:
            node_type = 'group'

        # Create node
        node_id = f'hts-{clean_code}' if clean_code else f'hts-group-{len(stack)}-{hash(desc) % 100000}'
        node = {
            'id': node_id,
            'code': htsno if htsno else '',
            'name': clean_desc,
            'type': node_type,
        }

        # For headings and groups, add children array
        if superior or node_type in ('heading', 'subheading', 'group'):
            node['children'] = []

        # Find parent based on indent level
        # Pop stack until we find a parent with lower indent
        while stack and stack[-1][0] >= indent:
            stack.pop()

        if stack:
            parent_node = stack[-1][1]
            if 'children' not in parent_node:
                parent_node['children'] = []
            parent_node['children'].append(node)
        else:
            # Top-level: attach to section
            if current_chapter is None and code_len >= 2:
                current_chapter = clean_code[:2]
            section = chapter_to_section.get(current_chapter or clean_code[:2], 'I')
            if section in section_nodes:
                section_nodes[section]['children'].append(node)

        # Push to stack if this node can have children
        if 'children' in node:
            stack.append((indent, node))

    # Clean empty children
    def clean_tree(nodes):
        for node in nodes:
            if 'children' in node:
                if len(node['children']) == 0:
                    del node['children']
                else:
                    clean_tree(node['children'])

    clean_tree(tree)

    # Build lookup
    lookup = {}

    def add_to_lookup(nodes, section='', section_name=''):
        for node in nodes:
            if node['type'] == 'section':
                section = node['code']
                section_name = node['name']
            code = node['code']
            if code and node['type'] != 'section':
                lookup[code] = {
                    'code': code,
                    'description': node['name'],
                    'section': section,
                    'sectionName': section_name,
                    'level': len(code.replace('.', '')),
                    'type': node['type']
                }
            if 'children' in node:
                add_to_lookup(node['children'], section, section_name)

    add_to_lookup(tree)

    write_json('hts-tree.json', tree)
    write_json('hts-lookup.json', lookup)
    print(f"  HTS: {len(tree)} sections, {len(lookup)} total codes in lookup")


# ─────────────────────────────────────────────
# 5. Canadian Customs Tariff
# ─────────────────────────────────────────────
def generate_canadian():
    print("\n=== Generating Canadian Tariff data ===")

    try:
        import pyodbc
    except ImportError:
        print("  ERROR: pyodbc not installed. Run: pip install pyodbc")
        return

    dbpath = os.path.join(RAW_DIR, '01-99-2025-0-eng.accdb')
    dbpath = os.path.abspath(dbpath)
    driver = 'Microsoft Access Driver (*.mdb, *.accdb)'
    conn = pyodbc.connect(f'Driver={{{driver}}};DBQ={dbpath};')
    cursor = conn.cursor()
    cursor.execute('SELECT TARIFF, DESC1, DESC2, DESC3 FROM TPHS ORDER BY TARIFF')

    rows = []
    for row in cursor.fetchall():
        tariff = str(row[0]).strip() if row[0] else ''
        desc1 = str(row[1]).strip() if row[1] else ''
        desc2 = str(row[2]).strip() if row[2] else ''
        desc3 = str(row[3]).strip() if row[3] else ''
        desc = desc1
        if desc2 and desc2 != 'None':
            desc += ' ' + desc2
        if desc3 and desc3 != 'None':
            desc += ' ' + desc3
        if tariff and desc:
            rows.append({'tariff': tariff, 'description': desc.strip()})

    conn.close()

    # Load HS sections
    hs_sections = {}
    with open(os.path.join(RAW_DIR, 'hs-sections.csv'), 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            hs_sections[row['section'].strip()] = row['name'].strip()

    chapter_to_section = {
        '01': 'I', '02': 'I', '03': 'I', '04': 'I', '05': 'I',
        '06': 'II', '07': 'II', '08': 'II', '09': 'II', '10': 'II', '11': 'II', '12': 'II', '13': 'II', '14': 'II',
        '15': 'III',
        '16': 'IV', '17': 'IV', '18': 'IV', '19': 'IV', '20': 'IV', '21': 'IV', '22': 'IV', '23': 'IV', '24': 'IV',
        '25': 'V', '26': 'V', '27': 'V',
        '28': 'VI', '29': 'VI', '30': 'VI', '31': 'VI', '32': 'VI', '33': 'VI', '34': 'VI', '35': 'VI', '36': 'VI', '37': 'VI', '38': 'VI',
        '39': 'VII', '40': 'VII',
        '41': 'VIII', '42': 'VIII', '43': 'VIII',
        '44': 'IX', '45': 'IX', '46': 'IX',
        '47': 'X', '48': 'X', '49': 'X',
        '50': 'XI', '51': 'XI', '52': 'XI', '53': 'XI', '54': 'XI', '55': 'XI', '56': 'XI', '57': 'XI', '58': 'XI', '59': 'XI', '60': 'XI', '61': 'XI', '62': 'XI', '63': 'XI',
        '64': 'XII', '65': 'XII', '66': 'XII', '67': 'XII',
        '68': 'XIII', '69': 'XIII', '70': 'XIII',
        '71': 'XIV',
        '72': 'XV', '73': 'XV', '74': 'XV', '75': 'XV', '76': 'XV', '78': 'XV', '79': 'XV', '80': 'XV', '81': 'XV', '82': 'XV', '83': 'XV',
        '84': 'XVI', '85': 'XVI',
        '86': 'XVII', '87': 'XVII', '88': 'XVII', '89': 'XVII',
        '90': 'XVIII', '91': 'XVIII', '92': 'XVIII',
        '93': 'XIX',
        '94': 'XX', '95': 'XX', '96': 'XX',
        '97': 'XXI',
        '98': 'XXII', '99': 'XXII',
    }

    # Build tree
    tree = []
    section_nodes = {}
    chapter_nodes_map = {}
    heading_nodes_map = {}
    subheading_nodes_map = {}
    tariff8_nodes_map = {}

    # Create section nodes
    for section_roman, section_name in hs_sections.items():
        node = {
            'id': f'ca-section-{section_roman}',
            'code': section_roman,
            'name': section_name.title() if section_name[0].islower() else section_name,
            'type': 'section',
            'children': []
        }
        section_nodes[section_roman] = node
        tree.append(node)

    # Add special section for ch 98-99
    if 'XXII' not in section_nodes:
        node = {
            'id': 'ca-section-XXII',
            'code': 'XXII',
            'name': 'Special Classification Provisions',
            'type': 'section',
            'children': []
        }
        section_nodes['XXII'] = node
        tree.append(node)

    for row in rows:
        tariff = row['tariff']
        desc = row['description']

        # Remove dots and determine level
        clean = tariff.replace('.', '')
        parts = tariff.split('.')

        # Determine type based on code structure
        if re.match(r'^\d{2}\.\d{2}$', tariff):
            # Heading: 01.01 -> 0101
            code = clean
            node = {
                'id': f'ca-{code}',
                'code': tariff,
                'name': desc,
                'type': 'heading',
                'children': []
            }
            heading_nodes_map[code] = node
            chapter = code[:2]
            section = chapter_to_section.get(chapter, 'I')
            # Create chapter node if not exists
            if chapter not in chapter_nodes_map:
                ch_node = {
                    'id': f'ca-{chapter}',
                    'code': chapter,
                    'name': f'Chapter {chapter}',
                    'type': 'chapter',
                    'children': []
                }
                chapter_nodes_map[chapter] = ch_node
                if section in section_nodes:
                    section_nodes[section]['children'].append(ch_node)
            chapter_nodes_map[chapter]['children'].append(node)

        elif re.match(r'^\d{4}\.\d$', tariff):
            # Intermediate grouping like 0101.2 -> group
            heading = clean[:4]
            node = {
                'id': f'ca-group-{clean}',
                'code': tariff,
                'name': desc,
                'type': 'group',
                'children': []
            }
            if heading in heading_nodes_map:
                heading_nodes_map[heading]['children'].append(node)
            # Store for child attachment
            subheading_nodes_map[clean] = node

        elif re.match(r'^\d{4}\.\d{2}\.\d{2}$', tariff):
            # 8-digit: 0101.21.00
            code_8 = clean
            code_6 = clean[:6]
            node = {
                'id': f'ca-{code_8}',
                'code': tariff,
                'name': desc,
                'type': 'subheading',
                'children': []
            }
            tariff8_nodes_map[code_8] = node
            # Find parent: try group first, then heading
            parent_found = False
            # Check for group parent (e.g., 01012 for 01012100)
            for gcode in sorted(subheading_nodes_map.keys(), key=len, reverse=True):
                if code_6.startswith(gcode):
                    subheading_nodes_map[gcode]['children'].append(node)
                    parent_found = True
                    break
            if not parent_found:
                heading = code_8[:4]
                if heading in heading_nodes_map:
                    heading_nodes_map[heading]['children'].append(node)

        elif re.match(r'^\d{4}\.\d{2}\.\d{2}\.\d{2}$', tariff):
            # 10-digit: 0101.21.00.00
            code_10 = clean
            code_8 = clean[:8]
            node = {
                'id': f'ca-{code_10}',
                'code': tariff,
                'name': desc,
                'type': 'tariff_item',
            }
            if code_8 in tariff8_nodes_map:
                tariff8_nodes_map[code_8]['children'].append(node)

    # Clean empty children
    def clean_tree(nodes):
        for node in nodes:
            if 'children' in node:
                if len(node['children']) == 0:
                    del node['children']
                else:
                    clean_tree(node['children'])

    clean_tree(tree)

    # Build lookup
    lookup = {}

    def add_to_lookup(nodes, section='', section_name=''):
        for node in nodes:
            if node['type'] == 'section':
                section = node['code']
                section_name = node['name']
            code = node['code']
            if code and node['type'] != 'section':
                clean = code.replace('.', '')
                lookup[code] = {
                    'code': code,
                    'description': node['name'],
                    'section': section,
                    'sectionName': section_name,
                    'level': len(clean),
                    'type': node['type']
                }
            if 'children' in node:
                add_to_lookup(node['children'], section, section_name)

    add_to_lookup(tree)

    write_json('ca-tree.json', tree)
    write_json('ca-lookup.json', lookup)
    print(f"  Canadian: {len(tree)} sections, {len(lookup)} total codes in lookup")


# ─────────────────────────────────────────────
# Generate concordance (stub - HS <-> CPC)
# ─────────────────────────────────────────────
def generate_concordance():
    print("\n=== Generating concordance data ===")

    csv_path = os.path.join(RAW_DIR, 'CPC21-HS2017.csv')
    if not os.path.exists(csv_path):
        print("  WARNING: CPC21-HS2017.csv not found in raw-data/. Generating empty concordance.")
        write_json('concordance.json', {'hsToCpc': {}, 'cpcToHs': {}, 'mappingInfo': {}})
        return

    hs_to_cpc = {}
    cpc_to_hs = {}

    with open(csv_path, 'r', encoding='utf-8-sig') as f:
        reader = csv.DictReader(f)
        for row in reader:
            hs_dotted = row['HS 2017'].strip()
            hs_partial = row['HS partial'].strip() == '1'
            cpc_code = row['CPC Ver. 2.1'].strip()
            cpc_partial = row['CPC partial'].strip() == '1'

            # Strip dots from HS code to match our pure-digit lookup keys
            hs_code = hs_dotted.replace('.', '')

            entry = {
                'hsPartial': hs_partial,
                'cpcPartial': cpc_partial,
            }

            hs_to_cpc.setdefault(hs_code, []).append({**entry, 'code': cpc_code})
            cpc_to_hs.setdefault(cpc_code, []).append({**entry, 'code': hs_code})

    # Build mappingInfo for tree node badge display
    mapping_info = {}
    for hs_code, mappings in hs_to_cpc.items():
        count = len(mappings)
        mapping_info[f'hs-{hs_code}'] = {'count': count, 'type': '1:1' if count == 1 else '1:N'}
    for cpc_code, mappings in cpc_to_hs.items():
        count = len(mappings)
        mapping_info[f'cpc-{cpc_code}'] = {'count': count, 'type': '1:1' if count == 1 else '1:N'}

    concordance = {
        'hsToCpc': hs_to_cpc,
        'cpcToHs': cpc_to_hs,
        'mappingInfo': mapping_info,
    }
    write_json('concordance.json', concordance)

    hs_count = len(hs_to_cpc)
    cpc_count = len(cpc_to_hs)
    total = sum(len(v) for v in hs_to_cpc.values())
    print(f"  Concordance: {total} mappings, {hs_count} HS codes, {cpc_count} CPC codes")


# ─────────────────────────────────────────────
# 7. UNSPSC (United Nations Standard Products and Services Code)
# ─────────────────────────────────────────────
def generate_unspsc():
    print("\n=== Generating UNSPSC data ===")

    # CSV columns: Segment, Segment Name, Family, Family Name, Class, Class Name, Commodity, Commodity Name
    # Each row is a commodity-level entry with parent hierarchy info
    segments = {}  # code -> name
    families = {}
    classes = {}
    commodities = {}

    with open(os.path.join(RAW_DIR, 'unspsc-codes.csv'), 'r', encoding='latin-1') as f:
        reader = csv.reader(f)
        next(reader)  # skip header
        for row in reader:
            if len(row) < 8:
                continue
            seg_code = row[0].strip()
            seg_name = row[1].strip()
            fam_code = row[2].strip()
            fam_name = row[3].strip()
            cls_code = row[4].strip()
            cls_name = row[5].strip()
            com_code = row[6].strip()
            com_name = row[7].strip()

            # Strip trailing zeros to get short codes for each level
            # Segment: first 2 digits, Family: first 4, Class: first 6, Commodity: all 8
            if seg_code and seg_name:
                segments[seg_code[:2]] = seg_name
            if fam_code and fam_name:
                families[fam_code[:4]] = fam_name
            if cls_code and cls_name:
                classes[cls_code[:6]] = cls_name
            if com_code and com_name:
                commodities[com_code[:8]] = com_name

    level_names = {2: 'segment', 4: 'family', 6: 'class', 8: 'commodity'}

    # Build tree top-down
    tree = []
    nodes_by_code = {}

    # Add segments (2-digit)
    for code in sorted(segments):
        node = {'id': f'unspsc-{code}', 'code': code, 'name': segments[code], 'type': 'segment', 'children': []}
        tree.append(node)
        nodes_by_code[code] = node

    # Add families (4-digit)
    for code in sorted(families):
        parent_code = code[:2]
        node = {'id': f'unspsc-{code}', 'code': code, 'name': families[code], 'type': 'family', 'children': []}
        nodes_by_code[code] = node
        if parent_code in nodes_by_code:
            nodes_by_code[parent_code]['children'].append(node)

    # Add classes (6-digit)
    for code in sorted(classes):
        parent_code = code[:4]
        node = {'id': f'unspsc-{code}', 'code': code, 'name': classes[code], 'type': 'class', 'children': []}
        nodes_by_code[code] = node
        if parent_code in nodes_by_code:
            nodes_by_code[parent_code]['children'].append(node)

    # Add commodities (8-digit) - leaf nodes, no children
    for code in sorted(commodities):
        parent_code = code[:6]
        node = {'id': f'unspsc-{code}', 'code': code, 'name': commodities[code], 'type': 'commodity'}
        nodes_by_code[code] = node
        if parent_code in nodes_by_code:
            nodes_by_code[parent_code]['children'].append(node)

    # Clean empty children
    def clean_tree(nodes):
        for node in nodes:
            if 'children' in node:
                if len(node['children']) == 0:
                    del node['children']
                else:
                    clean_tree(node['children'])

    clean_tree(tree)

    # Build lookup
    all_codes = {}
    all_codes.update({c: (segments[c], 'segment') for c in segments})
    all_codes.update({c: (families[c], 'family') for c in families})
    all_codes.update({c: (classes[c], 'class') for c in classes})
    all_codes.update({c: (commodities[c], 'commodity') for c in commodities})

    lookup = {}
    for code, (name, type_name) in all_codes.items():
        seg_code = code[:2]
        seg_name = segments.get(seg_code, '')
        lookup[code] = {
            'code': code,
            'description': name,
            'section': seg_code,
            'sectionName': seg_name,
            'level': len(code),
            'type': type_name,
        }

    write_json('unspsc-tree.json', tree)
    write_json('unspsc-lookup.json', lookup)
    print(f"  UNSPSC: {len(segments)} segments, {len(families)} families, {len(classes)} classes, {len(commodities)} commodities")
    print(f"  Total: {len(all_codes)} codes")


# ─────────────────────────────────────────────
# 8. UNSPSC ↔ HS Fuzzy Text Mapping
# ─────────────────────────────────────────────
def generate_unspsc_hs_fuzzy_mapping():
    print("\n=== Generating UNSPSC-HS fuzzy text mapping ===")

    # Load HS lookup (already generated)
    with open(os.path.join(OUT_DIR, 'hs-lookup.json'), 'r', encoding='utf-8') as f:
        hs_lookup = json.load(f)

    # Load UNSPSC lookup (already generated)
    with open(os.path.join(OUT_DIR, 'unspsc-lookup.json'), 'r', encoding='utf-8') as f:
        unspsc_lookup = json.load(f)

    # Get HS 6-digit subheadings and UNSPSC 8-digit commodities
    hs_items = {code: entry['description'] for code, entry in hs_lookup.items() if len(code) == 6}
    unspsc_items = {code: entry['description'] for code, entry in unspsc_lookup.items() if len(code) == 8}

    print(f"  HS subheadings: {len(hs_items)}, UNSPSC commodities: {len(unspsc_items)}")

    STOP_WORDS = {
        'a', 'an', 'and', 'or', 'the', 'of', 'for', 'to', 'in', 'on', 'with',
        'by', 'from', 'not', 'nor', 'but', 'at', 'as', 'is', 'are', 'was',
        'be', 'been', 'being', 'have', 'has', 'had', 'do', 'does', 'did',
        'other', 'than', 'its', 'it', 'their', 'this', 'that', 'these', 'those',
        'such', 'including', 'whether', 'also', 'etc', 'nes', 'nesoi',
    }

    def preprocess(text):
        text = text.lower()
        text = re.sub(r'[^\w\s]', ' ', text)
        words = text.split()
        return set(w for w in words if w not in STOP_WORDS and len(w) > 2)

    def jaccard(set1, set2):
        if not set1 or not set2:
            return 0.0
        intersection = len(set1 & set2)
        union = len(set1 | set2)
        return intersection / union if union > 0 else 0.0

    THRESHOLD = 0.3
    TOP_N = 3

    # Precompute HS word sets
    hs_word_sets = {code: preprocess(desc) for code, desc in hs_items.items()}

    unspsc_to_hs = {}
    hs_to_unspsc = {}
    count = 0

    for i, (unspsc_code, unspsc_desc) in enumerate(unspsc_items.items()):
        if i % 5000 == 0:
            print(f"    Processing {i}/{len(unspsc_items)} commodities...")

        unspsc_words = preprocess(unspsc_desc)
        if not unspsc_words:
            continue

        matches = []
        for hs_code, hs_words in hs_word_sets.items():
            sim = jaccard(unspsc_words, hs_words)
            if sim >= THRESHOLD:
                matches.append({'code': hs_code, 'similarity': round(sim, 3)})

        if matches:
            matches.sort(key=lambda x: x['similarity'], reverse=True)
            top = matches[:TOP_N]
            unspsc_to_hs[unspsc_code] = top
            count += len(top)

            for m in top:
                if m['code'] not in hs_to_unspsc:
                    hs_to_unspsc[m['code']] = []
                hs_to_unspsc[m['code']].append({
                    'code': unspsc_code,
                    'similarity': m['similarity'],
                })

    # Sort reverse mappings by similarity
    for hs_code in hs_to_unspsc:
        hs_to_unspsc[hs_code].sort(key=lambda x: x['similarity'], reverse=True)

    fuzzy_mapping = {
        'unspscToHs': unspsc_to_hs,
        'hsToUnspsc': hs_to_unspsc,
    }

    write_json('unspsc-hs-mapping.json', fuzzy_mapping)
    print(f"  Fuzzy mapping: {count} total matches")
    print(f"  UNSPSC codes matched: {len(unspsc_to_hs)}/{len(unspsc_items)} ({100*len(unspsc_to_hs)/len(unspsc_items):.1f}%)")
    print(f"  HS codes matched: {len(hs_to_unspsc)}/{len(hs_items)}")


# ─────────────────────────────────────────────
# 9. Taxonomy 1 (HTS goods + CPC services)
# ─────────────────────────────────────────────
def generate_taxonomy1():
    """Combine HTS (goods) with CPC service sections (5-9) into a single tree."""
    print("\n=== Generating Taxonomy 1 (HTS + CPC Services) ===")
    import copy

    # Load previously generated trees
    with open(os.path.join(OUT_DIR, 'hts-tree.json'), 'r', encoding='utf-8') as f:
        hts_tree = json.load(f)
    with open(os.path.join(OUT_DIR, 'cpc-tree.json'), 'r', encoding='utf-8') as f:
        cpc_tree = json.load(f)

    # Load previously generated lookups
    with open(os.path.join(OUT_DIR, 'hts-lookup.json'), 'r', encoding='utf-8') as f:
        hts_lookup = json.load(f)
    with open(os.path.join(OUT_DIR, 'cpc-lookup.json'), 'r', encoding='utf-8') as f:
        cpc_lookup = json.load(f)

    # Helper: re-prefix all node IDs in a tree
    def reprefix_tree(nodes, old_prefix, new_prefix):
        result = copy.deepcopy(nodes)
        def reprefix_node(node):
            if node['id'].startswith(old_prefix):
                node['id'] = new_prefix + node['id'][len(old_prefix):]
            if 'children' in node:
                for child in node['children']:
                    reprefix_node(child)
        for node in result:
            reprefix_node(node)
        return result

    # Re-prefix HTS tree: hts-* → t1-*
    t1_hts = reprefix_tree(hts_tree, 'hts-', 't1-')

    # Filter CPC to service sections (codes starting with 5-9) and re-prefix: cpc-* → t1-svc-*
    cpc_services = [s for s in cpc_tree if s['code'] in ('5', '6', '7', '8', '9')]
    t1_svc = reprefix_tree(cpc_services, 'cpc-', 't1-svc-')

    # Combined tree
    combined_tree = t1_hts + t1_svc

    # Build combined lookup
    combined_lookup = {}

    # HTS entries: keep original keys, add origin
    for key, entry in hts_lookup.items():
        combined_lookup[key] = {**entry, 'origin': 'hts'}

    # CPC service entries: prefix keys with SVC to avoid collisions, add origin + originalCode
    for key, entry in cpc_lookup.items():
        # Only include service codes (section 5-9)
        if entry.get('section', '') in ('5', '6', '7', '8', '9'):
            svc_key = f'SVC{key}'
            combined_lookup[svc_key] = {
                **entry,
                'code': entry['code'],  # keep original code for display
                'origin': 'cpc',
                'originalCode': key,    # original CPC code for concordance lookups
            }

    write_json('t1-tree.json', combined_tree)
    write_json('t1-lookup.json', combined_lookup)

    hts_count = sum(1 for v in combined_lookup.values() if v.get('origin') == 'hts')
    cpc_count = sum(1 for v in combined_lookup.values() if v.get('origin') == 'cpc')
    print(f"  Taxonomy 1: {len(t1_hts)} HTS sections + {len(t1_svc)} CPC service sections")
    print(f"  Lookup: {hts_count} HTS entries + {cpc_count} CPC service entries = {len(combined_lookup)} total")


# ─────────────────────────────────────────────
# 10. Taxonomy 2 (CPC backbone + HTS detail)
# ─────────────────────────────────────────────
def generate_taxonomy2():
    """Build T2: full CPC tree as backbone, with HTS tariff-line detail nested
    under CPC goods leaf nodes (sections 0-4) via the CPC↔HS concordance."""
    print("\n=== Generating Taxonomy 2 (CPC backbone + HTS detail) ===")
    import copy

    # Load previously generated data
    with open(os.path.join(OUT_DIR, 'cpc-tree.json'), 'r', encoding='utf-8') as f:
        cpc_tree = json.load(f)
    with open(os.path.join(OUT_DIR, 'cpc-lookup.json'), 'r', encoding='utf-8') as f:
        cpc_lookup = json.load(f)
    with open(os.path.join(OUT_DIR, 'hts-tree.json'), 'r', encoding='utf-8') as f:
        hts_tree = json.load(f)
    with open(os.path.join(OUT_DIR, 'hts-lookup.json'), 'r', encoding='utf-8') as f:
        hts_lookup = json.load(f)
    with open(os.path.join(OUT_DIR, 'concordance.json'), 'r', encoding='utf-8') as f:
        concordance = json.load(f)

    cpc_to_hs = concordance['cpcToHs']

    # Build HTS index: clean_code → node (with children) for all HTS nodes
    hts_index = {}
    def index_hts(nodes):
        for n in nodes:
            clean = n['code'].replace('.', '').replace(' ', '')
            if clean:
                hts_index[clean] = n
            if 'children' in n:
                index_hts(n['children'])
    index_hts(hts_tree)

    def extract_hts_for_hs(hs_code):
        """Extract HTS node(s) matching a 6-digit HS code. Returns list of deep-copied nodes."""
        # Case 1: Direct 6-digit subheading node exists
        if hs_code in hts_index:
            return [copy.deepcopy(hts_index[hs_code])]
        # Case 2: 8-digit +00 node (HS maps to single HTS tariff line)
        code_00 = hs_code + '00'
        if code_00 in hts_index:
            return [copy.deepcopy(hts_index[code_00])]
        # Case 3: Collect 8-digit HTS nodes whose code starts with this HS prefix
        matches = []
        for clean_code, node in hts_index.items():
            if clean_code.startswith(hs_code) and len(clean_code) == 8:
                matches.append(copy.deepcopy(node))
        if matches:
            return matches
        # Case 4: Try 10-digit nodes
        for clean_code, node in hts_index.items():
            if clean_code.startswith(hs_code) and len(clean_code) == 10:
                matches.append(copy.deepcopy(node))
        return matches

    # Re-prefix helper (same pattern as T1)
    def reprefix_tree(nodes, old_prefix, new_prefix):
        result = copy.deepcopy(nodes)
        def reprefix_node(node):
            if node['id'].startswith(old_prefix):
                node['id'] = new_prefix + node['id'][len(old_prefix):]
            if 'children' in node:
                for child in node['children']:
                    reprefix_node(child)
        for node in result:
            reprefix_node(node)
        return result

    def reprefix_node_inplace(node, old_prefix, new_prefix):
        if node['id'].startswith(old_prefix):
            node['id'] = new_prefix + node['id'][len(old_prefix):]
        if 'children' in node:
            for child in node['children']:
                reprefix_node_inplace(child, old_prefix, new_prefix)

    # Deep copy full CPC tree, re-prefix cpc-* → t2-*
    t2_tree = reprefix_tree(cpc_tree, 'cpc-', 't2-')

    # Track used IDs to handle duplicates (some HS codes map to multiple CPC parents)
    used_ids = set()

    def make_unique_ids(node):
        """Ensure node and all descendants have unique IDs, appending -dN for dupes."""
        if node['id'] in used_ids:
            n = 2
            while f"{node['id']}-d{n}" in used_ids:
                n += 1
            node['id'] = f"{node['id']}-d{n}"
        used_ids.add(node['id'])
        if 'children' in node:
            for child in node['children']:
                make_unique_ids(child)

    # Walk the T2 tree to collect all CPC backbone IDs first
    def collect_ids(nodes):
        for n in nodes:
            used_ids.add(n['id'])
            if 'children' in n:
                collect_ids(n['children'])
    collect_ids(t2_tree)

    # Determine which top-level sections are goods (codes 0-4)
    goods_section_codes = {'0', '1', '2', '3', '4'}

    hts_nodes_added = 0

    def attach_hts_detail(nodes, in_goods_section=False):
        nonlocal hts_nodes_added
        for node in nodes:
            # Determine if we're inside a goods section
            original_cpc_code = node['id'].replace('t2-', '')
            is_goods = in_goods_section or (len(original_cpc_code) == 1 and original_cpc_code in goods_section_codes)

            if 'children' in node:
                # Non-leaf: recurse
                attach_hts_detail(node['children'], is_goods)
            elif is_goods:
                # Leaf in goods section: try to attach HTS detail
                hs_mappings = cpc_to_hs.get(original_cpc_code, [])
                if hs_mappings:
                    hts_children = []
                    for mapping in hs_mappings:
                        hs_code = mapping['code']
                        extracted = extract_hts_for_hs(hs_code)
                        for ext_node in extracted:
                            reprefix_node_inplace(ext_node, 'hts-', 't2-hts-')
                            make_unique_ids(ext_node)
                            hts_children.append(ext_node)
                            hts_nodes_added += 1
                    if hts_children:
                        node['children'] = hts_children

    attach_hts_detail(t2_tree)

    # Build combined lookup
    combined_lookup = {}

    # CPC backbone entries: keep original keys, add origin
    for key, entry in cpc_lookup.items():
        combined_lookup[key] = {**entry, 'origin': 'cpc'}

    # HTS detail entries: walk T2 tree to find all t2-hts-* nodes
    def add_hts_to_lookup(nodes, section_code='', section_name=''):
        for node in nodes:
            if node['id'].startswith('t2-hts-'):
                clean_code = node['code'].replace('.', '').replace(' ', '')
                hts_key = f'HTS{clean_code}'
                # For duplicate nodes (e.g. t2-hts-010121-d2), use the base code
                if hts_key not in combined_lookup:
                    combined_lookup[hts_key] = {
                        'code': node['code'],
                        'description': node['name'],
                        'section': section_code,
                        'sectionName': section_name,
                        'level': len(clean_code),
                        'type': node.get('type', 'detail'),
                        'origin': 'hts',
                        'originalCode': clean_code,
                    }
            else:
                # CPC backbone node - track section context
                cpc_code = node['id'].replace('t2-', '')
                if len(cpc_code) == 1:
                    section_code = cpc_code
                    section_name = node['name']
            if 'children' in node:
                add_hts_to_lookup(node['children'], section_code, section_name)

    add_hts_to_lookup(t2_tree)

    write_json('t2-tree.json', t2_tree)
    write_json('t2-lookup.json', combined_lookup)

    cpc_count = sum(1 for v in combined_lookup.values() if v.get('origin') == 'cpc')
    hts_count = sum(1 for v in combined_lookup.values() if v.get('origin') == 'hts')
    print(f"  Taxonomy 2: CPC backbone ({cpc_count} entries) + HTS detail ({hts_count} entries)")
    print(f"  HTS nodes attached to CPC leaves: {hts_nodes_added}")
    print(f"  Total lookup entries: {len(combined_lookup)}")


if __name__ == '__main__':
    print("Generating taxonomy data...")
    generate_hs()
    generate_cpc()
    generate_cn()
    generate_hts()
    generate_canadian()
    generate_concordance()
    generate_unspsc()
    generate_unspsc_hs_fuzzy_mapping()
    generate_taxonomy1()
    generate_taxonomy2()
    print("\nDone!")
