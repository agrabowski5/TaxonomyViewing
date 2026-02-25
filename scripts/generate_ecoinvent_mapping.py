"""
Generate ecoinvent mapping JSON from the Database Overview Excel file.
Extracts product-to-CPC, product-to-HS, and product-to-ISIC mappings
from all allocation sheets and the Intermediate Exchanges sheet.
"""
import json
import os
from collections import defaultdict


def generate_ecoinvent_mapping():
    import openpyxl

    excel_path = os.path.join(os.path.dirname(__file__), '..', 'raw-data',
                              'Database-Overview-for-ecoinvent-v3.10.xlsx')
    output_path = os.path.join(os.path.dirname(__file__), '..', 'app', 'public', 'data',
                               'ecoinvent-mapping.json')

    print("Loading ecoinvent database overview...")
    wb = openpyxl.load_workbook(excel_path, read_only=True, data_only=True)

    # Column layouts differ between sheets (1-indexed):
    # Cut-Off/EN15804/APOS: Product=12, CPC=13, HS=14, ISIC=9
    # Consequential: Product=13, CPC=14, HS=15, ISIC=10 (extra "Technology Level" col)
    # Undefined: Product=10, CPC=11, HS=12, ISIC=6
    sheets_config = {
        'Cut-Off AO':       {'product': 12, 'cpc': 13, 'hs': 14, 'isic': 9},
        'EN15804 AO':       {'product': 12, 'cpc': 13, 'hs': 14, 'isic': 9},
        'APOS AO':          {'product': 12, 'cpc': 13, 'hs': 14, 'isic': 9},
        'Consequential AO': {'product': 13, 'cpc': 14, 'hs': 15, 'isic': 10},
        'Undefined AO':     {'product': 10, 'cpc': 11, 'hs': 12, 'isic': 6},
    }

    product_cpc = {}       # product name -> CPC code
    product_hs = {}        # product name -> HS code
    product_isic = {}      # product name -> ISIC code
    product_cpc_desc = {}
    product_hs_desc = {}
    product_isic_desc = {}

    def parse_classification(raw):
        """Parse 'CODE: Description' format, return (code, description) or (None, None)."""
        if not raw or str(raw).strip() in ('', 'None', 'not available'):
            return None, None
        s = str(raw).strip()
        if ':' not in s:
            return None, None
        parts = s.split(':', 1)
        code = parts[0].strip()
        desc = parts[1].strip() if len(parts) > 1 else ''
        if not code:
            return None, None
        return code, desc

    total_rows = 0

    # Process all allocation sheets
    for sheet_name, cols in sheets_config.items():
        if sheet_name not in wb.sheetnames:
            print(f"  Skipping {sheet_name} (not found)")
            continue

        ws = wb[sheet_name]
        sheet_rows = 0

        for row in ws.iter_rows(min_row=2, values_only=True):
            product = str(row[cols['product'] - 1]).strip() if len(row) >= cols['product'] and row[cols['product'] - 1] else ''
            if not product or product == 'None':
                continue

            cpc_raw = row[cols['cpc'] - 1] if len(row) >= cols['cpc'] else None
            hs_raw = row[cols['hs'] - 1] if len(row) >= cols['hs'] else None
            isic_raw = row[cols['isic'] - 1] if len(row) >= cols['isic'] else None

            cpc_code, cpc_desc = parse_classification(cpc_raw)
            hs_code, hs_desc = parse_classification(hs_raw)
            isic_code, isic_desc = parse_classification(isic_raw)

            if cpc_code and product not in product_cpc:
                product_cpc[product] = cpc_code
                product_cpc_desc[product] = cpc_desc

            if hs_code and product not in product_hs:
                product_hs[product] = hs_code
                product_hs_desc[product] = hs_desc

            if isic_code and product not in product_isic:
                product_isic[product] = isic_code
                product_isic_desc[product] = isic_desc

            sheet_rows += 1

        total_rows += sheet_rows
        print(f"  {sheet_name}: {sheet_rows} rows")

    # Also process Intermediate Exchanges for additional products with CPC
    if 'Intermediate Exchanges' in wb.sheetnames:
        ws = wb['Intermediate Exchanges']
        ie_rows = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            name = str(row[1]).strip() if len(row) >= 2 and row[1] else ''
            if not name or name == 'None':
                continue
            cpc_raw = row[6] if len(row) >= 7 else None  # Column 7: CPC Classification
            cpc_code, cpc_desc = parse_classification(cpc_raw)
            if cpc_code and name not in product_cpc:
                product_cpc[name] = cpc_code
                product_cpc_desc[name] = cpc_desc
            ie_rows += 1
        total_rows += ie_rows
        print(f"  Intermediate Exchanges: {ie_rows} rows")

    wb.close()

    all_products = set(product_cpc.keys()) | set(product_hs.keys()) | set(product_isic.keys())
    print(f"\n  Total rows processed: {total_rows}")
    print(f"  Unique products: {len(all_products)}")
    print(f"  Products with CPC: {len(product_cpc)}")
    print(f"  Products with HS: {len(product_hs)}")
    print(f"  Products with ISIC: {len(product_isic)}")

    # Build reverse maps: code -> list of products
    cpc_to_products = defaultdict(list)
    hs_to_products = defaultdict(list)
    isic_to_products = defaultdict(list)

    for product, code in product_cpc.items():
        cpc_to_products[code].append(product)
    for product, code in product_hs.items():
        hs_to_products[code].append(product)
    for product, code in product_isic.items():
        isic_to_products[code].append(product)

    # Build output mappings
    def build_mappings(code_to_products):
        mappings = {}
        for code, products in sorted(code_to_products.items()):
            n = len(products)
            mappings[code] = {
                "products": sorted(products),
                "count": n,
                "mappingType": "1:1" if n == 1 else f"{n}:1",
            }
        return mappings

    cpc_mappings = build_mappings(cpc_to_products)
    hs_mappings = build_mappings(hs_to_products)
    isic_mappings = build_mappings(isic_to_products)

    # Build ancestor coverage for tree highlighting
    # CPC: 5-digit -> 4-digit -> 3-digit -> 2-digit -> 1-digit (section)
    # HS: 6-digit -> 4-digit -> 2-digit
    # ISIC: 4-digit -> 3-digit -> 2-digit
    def build_ancestors(codes):
        ancestors = set()
        for code in codes:
            for length in range(1, len(code)):
                ancestors.add(code[:length])
        return ancestors

    cpc_ancestors = build_ancestors(cpc_to_products.keys())
    hs_ancestors = build_ancestors(hs_to_products.keys())
    isic_ancestors = build_ancestors(isic_to_products.keys())

    result = {
        "cpc": cpc_mappings,
        "hs": hs_mappings,
        "isic": isic_mappings,
        "cpcAncestors": sorted(cpc_ancestors),
        "hsAncestors": sorted(hs_ancestors),
        "isicAncestors": sorted(isic_ancestors),
        "stats": {
            "totalProducts": len(all_products),
            "productsWithCpc": len(product_cpc),
            "productsWithHs": len(product_hs),
            "productsWithIsic": len(product_isic),
            "uniqueCpcCodes": len(cpc_to_products),
            "uniqueHsCodes": len(hs_to_products),
            "uniqueIsicCodes": len(isic_to_products),
        }
    }

    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(result, f, ensure_ascii=False)

    file_size = os.path.getsize(output_path) / 1024 / 1024
    print(f"\n  Output: {output_path} ({file_size:.1f} MB)")
    print(f"  Stats: {result['stats']}")
    print(f"  CPC codes: {len(cpc_mappings)}, ancestors: {len(cpc_ancestors)}")
    print(f"  HS codes: {len(hs_mappings)}, ancestors: {len(hs_ancestors)}")
    print(f"  ISIC codes: {len(isic_mappings)}, ancestors: {len(isic_ancestors)}")


if __name__ == '__main__':
    generate_ecoinvent_mapping()
